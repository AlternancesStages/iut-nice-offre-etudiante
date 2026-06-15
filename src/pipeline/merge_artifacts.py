# merge_artifacts.py

import os
import pandas as pd
from openpyxl import load_workbook

OUTPUT_DIR = "exports"

DOMAIN_FILES = [
    "offres_sd.xlsx",
    "offres_reseaux_telecom.xlsx",
    "offres_qlio.xlsx",
    "offres_cs.xlsx",
    "offres_gea.xlsx",
    "offres_infocom.xlsx",
    "offres_info.xlsx",
    "offres_tc.xlsx",
    "offres_geii.xlsx",
]

SOURCE_ARTIFACTS = [
    "temp/legal_sources",
    "temp/jobhive",
]

def artifact_directories():
    """Repere les dossiers d'artifacts telecharges par GitHub Actions."""
    directories = []

    for artifact_dir in SOURCE_ARTIFACTS:
        if not os.path.exists(artifact_dir):
            print(f"Dossier artifact absent : {artifact_dir}")
            continue

        has_workbooks = any(
            os.path.exists(os.path.join(artifact_dir, filename))
            for filename in DOMAIN_FILES
        )

        if has_workbooks:
            directories.append(artifact_dir)

        for child in sorted(os.listdir(artifact_dir)):
            child_path = os.path.join(artifact_dir, child)

            if os.path.isdir(child_path):
                directories.append(child_path)

    return directories

def ensure_output_files():
    """Cree les classeurs de sortie avant d'y remplacer les feuilles sources."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for filename in DOMAIN_FILES:
        output_path = os.path.join(OUTPUT_DIR, filename)

        if not os.path.exists(output_path):
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                pd.DataFrame().to_excel(
                    writer,
                    index=False,
                    sheet_name="INIT"
                )

def copy_sheets_from_artifact(artifact_dir, filename):
    """Copie les feuilles d'un artifact source vers le classeur de sortie."""
    source_file = os.path.join(artifact_dir, filename)
    output_file = os.path.join(OUTPUT_DIR, filename)

    if not os.path.exists(source_file):
        print(f"Fichier absent : {source_file}")
        return

    try:
        sheets = pd.read_excel(source_file, sheet_name=None)
    except Exception as e:
        print(f"Erreur lecture {source_file} : {e}")
        return

    for sheet_name, df in sheets.items():
        if sheet_name == "INIT":
            continue

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name=sheet_name
            )

        print(f"Feuille copiée : {sheet_name} -> {output_file}")

def remove_init_sheets():
    """Supprime la feuille temporaire INIT quand au moins une vraie feuille existe."""
    for filename in DOMAIN_FILES:
        output_file = os.path.join(OUTPUT_DIR, filename)

        if not os.path.exists(output_file):
            continue

        wb = load_workbook(output_file)

        if "INIT" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["INIT"]
            wb.save(output_file)
            print(f"Feuille INIT supprimée : {output_file}")

def main():
    """Fusionne les artifacts produits par les jobs paralleles de collecte."""
    ensure_output_files()

    for artifact_dir in artifact_directories():
        for filename in DOMAIN_FILES:
            copy_sheets_from_artifact(artifact_dir, filename)

    remove_init_sheets()

if __name__ == "__main__":
    main()
